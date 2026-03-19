"""
Export Service for generating reports in various formats.
"""

import os
import json
from typing import List, Any
from datetime import datetime
import uuid

# Try to import pandas and openpyxl
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


EXPORT_DIR = os.environ.get("EXPORT_DIR", "./exports")
os.makedirs(EXPORT_DIR, exist_ok=True)


class ExportService:
    """Service for exporting trip data to various formats."""
    
    async def export_trip_to_excel(self, trip: Any, receipts: List[Any]) -> str:
        """
        Export trip data to Excel format with summary and charts.
        
        Args:
            trip: Trip object
            receipts: List of Receipt objects
        
        Returns:
            Path to the generated Excel file
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")
        
        wb = Workbook()
        
        # Create Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Trip Summary"
        
        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Trip Info Header
        ws_summary["A1"] = "Trip Information"
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary.merge_cells("A1:B1")
        
        # Trip details
        trip_info = [
            ("Trip Name", trip.name),
            ("Description", trip.description or ""),
            ("Start Date", str(trip.start_date) if trip.start_date else ""),
            ("End Date", str(trip.end_date) if trip.end_date else ""),
            ("Budget", f"{trip.currency} {trip.budget}" if trip.budget else ""),
            ("Currency", trip.currency),
            ("Total Receipts", len(receipts)),
            ("Total Spent", f"{trip.currency} {sum(r.total_amount for r in receipts):.2f}"),
        ]
        
        for i, (label, value) in enumerate(trip_info, start=3):
            ws_summary[f"A{i}"] = label
            ws_summary[f"A{i}"].font = Font(bold=True)
            ws_summary[f"B{i}"] = value
        
        # Category breakdown
        ws_summary["A13"] = "Spending by Category"
        ws_summary["A13"].font = Font(bold=True, size=12)
        ws_summary.merge_cells("A13:B13")
        
        category_spending = {}
        for receipt in receipts:
            cat = receipt.category or "Uncategorized"
            category_spending[cat] = category_spending.get(cat, 0) + receipt.total_amount
        
        ws_summary["A15"] = "Category"
        ws_summary["B15"] = "Amount"
        ws_summary["A15"].font = header_font
        ws_summary["B15"].font = header_font
        ws_summary["A15"].fill = header_fill
        ws_summary["B15"].fill = header_fill
        
        for i, (cat, amount) in enumerate(sorted(category_spending.items()), start=16):
            ws_summary[f"A{i}"] = cat
            ws_summary[f"B{i}"] = f"{trip.currency} {amount:.2f}"
        
        # Adjust column widths
        ws_summary.column_dimensions["A"].width = 20
        ws_summary.column_dimensions["B"].width = 25
        
        # Create Receipts sheet
        ws_receipts = wb.create_sheet("Receipts")
        
        headers = [
            "Date", "Merchant", "Category", "Amount", "Currency",
            "Address", "Notes"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws_receipts.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row, receipt in enumerate(receipts, start=2):
            ws_receipts.cell(row=row, column=1, value=str(receipt.purchase_date) if receipt.purchase_date else "")
            ws_receipts.cell(row=row, column=2, value=receipt.merchant_name or "")
            ws_receipts.cell(row=row, column=3, value=receipt.category or "")
            ws_receipts.cell(row=row, column=4, value=receipt.total_amount)
            ws_receipts.cell(row=row, column=5, value=receipt.currency)
            ws_receipts.cell(row=row, column=6, value=receipt.address or "")
            ws_receipts.cell(row=row, column=7, value=receipt.notes or "")
            
            for col in range(1, 8):
                ws_receipts.cell(row=row, column=col).border = border
        
        # Adjust column widths
        column_widths = [15, 25, 15, 12, 10, 30, 30]
        for i, width in enumerate(column_widths, start=1):
            ws_receipts.column_dimensions[chr(64 + i)].width = width
        
        # Create Locations sheet (for mapping)
        ws_locations = wb.create_sheet("Locations")
        
        loc_headers = ["Merchant", "Address", "Latitude", "Longitude", "Amount"]
        for col, header in enumerate(loc_headers, start=1):
            cell = ws_locations.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row = 2
        for receipt in receipts:
            if receipt.latitude and receipt.longitude:
                ws_locations.cell(row=row, column=1, value=receipt.merchant_name or "")
                ws_locations.cell(row=row, column=2, value=receipt.address or "")
                ws_locations.cell(row=row, column=3, value=receipt.latitude)
                ws_locations.cell(row=row, column=4, value=receipt.longitude)
                ws_locations.cell(row=row, column=5, value=receipt.total_amount)
                row += 1
        
        # Save file
        filename = f"trip_{trip.id}_{uuid.uuid4().hex[:8]}.xlsx"
        file_path = os.path.join(EXPORT_DIR, filename)
        wb.save(file_path)
        
        return file_path
    
    async def export_trip_to_csv(self, trip: Any, receipts: List[Any]) -> str:
        """
        Export trip data to CSV format.
        
        Args:
            trip: Trip object
            receipts: List of Receipt objects
        
        Returns:
            Path to the generated CSV file
        """
        import csv
        
        filename = f"trip_{trip.id}_{uuid.uuid4().hex[:8]}.csv"
        file_path = os.path.join(EXPORT_DIR, filename)
        
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            
            # Header
            writer.writerow([
                "Date", "Merchant", "Category", "Amount", "Currency",
                "Address", "Latitude", "Longitude", "Notes"
            ])
            
            # Data rows
            for receipt in receipts:
                writer.writerow([
                    str(receipt.purchase_date) if receipt.purchase_date else "",
                    receipt.merchant_name or "",
                    receipt.category or "",
                    receipt.total_amount,
                    receipt.currency,
                    receipt.address or "",
                    receipt.latitude or "",
                    receipt.longitude or "",
                    receipt.notes or ""
                ])
        
        return file_path
    
    async def export_trip_to_json(self, trip: Any, receipts: List[Any]) -> dict:
        """
        Export trip data to JSON format.
        
        Args:
            trip: Trip object
            receipts: List of Receipt objects
        
        Returns:
            Dictionary containing trip data
        """
        # Calculate statistics
        total_spent = sum(r.total_amount for r in receipts)
        
        category_spending = {}
        for receipt in receipts:
            cat = receipt.category or "Uncategorized"
            category_spending[cat] = category_spending.get(cat, 0) + receipt.total_amount
        
        return {
            "trip": {
                "id": trip.id,
                "name": trip.name,
                "description": trip.description,
                "start_date": str(trip.start_date) if trip.start_date else None,
                "end_date": str(trip.end_date) if trip.end_date else None,
                "budget": trip.budget,
                "currency": trip.currency,
                "created_at": trip.created_at.isoformat() if trip.created_at else None,
            },
            "statistics": {
                "total_spent": total_spent,
                "receipt_count": len(receipts),
                "remaining_budget": (trip.budget or 0) - total_spent,
                "category_breakdown": category_spending
            },
            "receipts": [
                {
                    "id": r.id,
                    "merchant_name": r.merchant_name,
                    "total_amount": r.total_amount,
                    "currency": r.currency,
                    "category": r.category,
                    "purchase_date": r.purchase_date.isoformat() if r.purchase_date else None,
                    "items": r.items,
                    "address": r.address,
                    "latitude": r.latitude,
                    "longitude": r.longitude,
                    "notes": r.notes
                }
                for r in receipts
            ]
        }
